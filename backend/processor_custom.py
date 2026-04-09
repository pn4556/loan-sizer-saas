"""
Custom Loan Processor for Multi and Mixed-Use Term Sizer
Specifically configured for: Multi and Mixed-Use Term Sizer 3.18.2026 1.xlsx
"""

import os
import subprocess
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from sizer_config import (
    SIZER_CELL_MAPPINGS,
    PROGRAM_SHEETS,
    PROGRAM_STATUS_CELLS,
    REQUIRED_FIELDS,
    VALIDATION_RULES,
    US_STATES
)


@dataclass
class LoanApplication:
    """Loan application data structure"""
    # Property Details (required)
    units: int
    address: str
    city: str
    state: str
    zip_code: str
    
    # Financial (required)
    estimated_value: float
    purchase_price: float
    loan_amount: float
    note_type: str
    
    # Credit (3 scores - will calculate middle) (required)
    credit_score_1: int
    credit_score_2: int
    credit_score_3: int
    
    # Optional fields (all have defaults)
    points_to_lender: float = 0.0
    unit_size: Optional[str] = None
    property_type: str = "Multifamily"
    asset_class: str = "C"
    occupancy: float = 1.0
    rent_stabilized: str = "No"
    square_footage: Optional[float] = None
    
    @property
    def credit_score_middle(self) -> int:
        """Return middle credit score (standard lending practice)"""
        scores = sorted([self.credit_score_1, self.credit_score_2, self.credit_score_3])
        return scores[1]
    
    @property
    def ltv_ratio(self) -> float:
        """Calculate Loan-to-Value ratio"""
        return (self.loan_amount / self.estimated_value) * 100
    
    @property
    def full_address(self) -> str:
        """Return formatted full address"""
        return f"{self.address}, {self.city}, {self.state} {self.zip_code}"
    
    def validate(self) -> Tuple[bool, List[str]]:
        """Validate application data"""
        errors = []
        
        # Check required fields
        for field in REQUIRED_FIELDS:
            value = getattr(self, field, None)
            if value is None or value == '':
                errors.append(f"Missing required field: {field}")
        
        # Validate state
        if self.state and self.state.upper() not in US_STATES:
            errors.append(f"Invalid state: {self.state}")
        
        # Validate credit scores
        for i, score in enumerate([self.credit_score_1, self.credit_score_2, self.credit_score_3], 1):
            if not 300 <= score <= 850:
                errors.append(f"Credit score {i} must be between 300-850")
        
        # Validate loan amount <= estimated value
        if self.loan_amount > self.estimated_value:
            errors.append("Loan amount cannot exceed estimated value")
        
        return len(errors) == 0, errors


@dataclass
class ProgramResult:
    """Result from a single program evaluation"""
    name: str
    status: str  # "PASS", "FAIL", or "REVIEW"
    max_loan_amount: Optional[float]
    interest_rate: Optional[float]
    dscr: Optional[float] = None
    ltv_max: Optional[float] = None
    reason: Optional[str] = None
    details: Optional[Dict] = None


@dataclass
class ProcessingResult:
    """Complete processing result"""
    application: LoanApplication
    programs: List[ProgramResult]
    overall_decision: str  # "APPROVE", "REJECT", "REVIEW"
    decision_reason: str
    output_file: str
    processing_time: float
    timestamp: datetime


class SizerProcessor:
    """Processor for the Multi and Mixed-Use Term Sizer Excel file"""
    
    def __init__(self, template_path: str):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
    
    def process_application(
        self, 
        application: LoanApplication, 
        interest_rate: float,
        output_dir: Optional[str] = None
    ) -> ProcessingResult:
        """
        Process a loan application through the sizer
        
        Args:
            application: LoanApplication object with all fields
            interest_rate: Daily interest rate
            output_dir: Directory to save output file (default: temp)
        
        Returns:
            ProcessingResult with all evaluation data
        """
        import time
        start_time = time.time()
        
        # Validate application
        is_valid, errors = application.validate()
        if not is_valid:
            raise ValueError(f"Validation failed: {'; '.join(errors)}")
        
        # Create output file
        if output_dir is None:
            output_dir = tempfile.gettempdir()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"sizer_{application.address.replace(' ', '_')[:20]}_{timestamp}.xlsx"
        output_path = Path(output_dir) / output_filename
        
        # Load and populate workbook
        wb = load_workbook(self.template_path, data_only=False)
        
        try:
            # Get SIZER sheet
            if 'SIZER' in wb.sheetnames:
                sizer_sheet = wb['SIZER']
            else:
                sizer_sheet = wb.active
            
            # Populate input fields
            self._populate_sizer_sheet(sizer_sheet, application, interest_rate)
            
            # Save workbook (formulas will recalculate on open)
            wb.save(output_path)
            
            # Recalculate using LibreOffice if available
            self._recalculate_workbook(output_path)
            
            # Reload to get calculated values
            wb_calc = load_workbook(output_path, data_only=True)
            if 'SIZER' in wb_calc.sheetnames:
                sizer_calc = wb_calc['SIZER']
            else:
                sizer_calc = wb_calc.active
            
            # Evaluate programs
            programs = self._evaluate_programs(sizer_calc, application)
            
            # Make overall decision
            decision, reason = self._make_decision(programs)
            
            processing_time = time.time() - start_time
            
            return ProcessingResult(
                application=application,
                programs=programs,
                overall_decision=decision,
                decision_reason=reason,
                output_file=str(output_path),
                processing_time=processing_time,
                timestamp=datetime.now()
            )
            
        finally:
            wb.close()
    
    def _populate_sizer_sheet(self, sheet, application: LoanApplication, interest_rate: float):
        """Populate the SIZER sheet with application data"""
        
        # Map application fields to cells
        field_values = {
            'units': application.units,
            'address': application.address,
            'city': application.city,
            'state': application.state,
            'zip_code': application.zip_code,
            'estimated_value': application.estimated_value,
            'purchase_price': application.purchase_price,
            'loan_amount': application.loan_amount,
            'note_type': application.note_type,
            'points_to_lender': application.points_to_lender,
            'credit_score': application.credit_score_middle,
            'interest_rate': interest_rate,
            'property_type': application.property_type,
            'asset_class': application.asset_class,
            'occupancy': application.occupancy,
            'rent_stabilized': application.rent_stabilized,
        }
        
        # Calculate unit size if square footage provided
        if application.square_footage and application.units:
            field_values['unit_size'] = application.square_footage / application.units
            field_values['square_footage'] = application.square_footage
        
        # Write to cells
        for cell_ref, field_name in SIZER_CELL_MAPPINGS.items():
            if field_name in field_values:
                value = field_values[field_name]
                sheet[cell_ref] = value
    
    def _recalculate_workbook(self, file_path: str):
        """Recalculate Excel formulas using LibreOffice"""
        try:
            # Try LibreOffice command line recalculation
            result = subprocess.run(
                ['libreoffice', '--headless', '--calc', '--nologo', 
                 '--norestore', '--accept=socket,host=localhost,port=2002;urp;',
                 file_path],
                capture_output=True,
                timeout=30
            )
        except (subprocess.TimeoutExpired, FileNotFoundError):
            # LibreOffice not available - formulas will recalc on open
            pass
    
    def _evaluate_programs(self, sheet, application: LoanApplication) -> List[ProgramResult]:
        """Evaluate all loan programs and return results"""
        
        programs = []
        
        # Read calculated values from sheet
        try:
            final_dscr = sheet['E40'].value if sheet['E40'].value else 0
            ltv_ratio = sheet['E45'].value if sheet['E45'].value else application.ltv_ratio
            state_status = sheet['E10'].value if sheet['E10'].value else "--"
        except:
            final_dscr = 0
            ltv_ratio = application.ltv_ratio
            state_status = "--"
        
        # Program 1: Insurance Program
        # Criteria: LTV <= 80%, Credit >= 640, State eligible
        if (application.ltv_ratio <= 80 and 
            application.credit_score_middle >= 640 and 
            state_status != "Fail"):
            
            programs.append(ProgramResult(
                program_name="Insurance Program",
                status="PASS",
                decision="APPROVE",
                interest_rate=7.5,
                reason="Meets all requirements"
            ))
        else:
            fail_reasons = []
            if application.ltv_ratio > 80:
                fail_reasons.append(f"LTV {application.ltv_ratio:.1f}% > 80%")
            if application.credit_score_middle < 640:
                fail_reasons.append(f"Credit {application.credit_score_middle} < 640")
            if state_status == "Fail":
                fail_reasons.append("State not eligible")
            
            programs.append(ProgramResult(
                program_name="Insurance Program",
                status="FAIL",
                decision="DECLINE",
                interest_rate=0.0,
                reason="; ".join(fail_reasons) if fail_reasons else "Does not meet requirements"
            ))
        
        # Program 2: Short Term Sale
        # Criteria: LTV <= 75%, Credit >= 680
        if application.ltv_ratio <= 75 and application.credit_score_middle >= 680:
            programs.append(ProgramResult(
                program_name="Short Term Sale",
                status="PASS",
                decision="APPROVE",
                interest_rate=7.5,
                reason="Meets Short Term Sale requirements"
            ))
        else:
            programs.append(ProgramResult(
                program_name="Short Term Sale",
                status="FAIL",
                decision="DECLINE",
                interest_rate=0.0,
                reason="LTV or credit below requirements"
            ))
        
        # Program 3: Deephaven
        # Criteria: LTV <= 70%, Credit >= 720
        if application.ltv_ratio <= 70 and application.credit_score_middle >= 720:
            programs.append(ProgramResult(
                program_name="Deephaven",
                status="PASS",
                decision="APPROVE",
                interest_rate=7.5,
                reason="Meets Deephaven requirements"
            ))
        else:
            programs.append(ProgramResult(
                program_name="Deephaven",
                status="FAIL",
                decision="DECLINE",
                interest_rate=0.0,
                reason="LTV or credit below requirements"
            ))
        
        # Additional programs can be added based on specific criteria
        
        return programs
    
    def _make_decision(self, programs: List[ProgramResult]) -> Tuple[str, str]:
        """Make overall approval decision based on program results"""
        
        pass_count = sum(1 for p in programs if p.status == "PASS")
        fail_count = sum(1 for p in programs if p.status == "FAIL")
        
        if pass_count >= 2:
            return "APPROVE", f"Qualified for {pass_count} of {len(programs)} programs"
        elif pass_count == 1:
            return "REVIEW", "Marginal qualification - manual review recommended"
        else:
            return "REJECT", "Does not meet minimum program requirements"


# Example usage
if __name__ == "__main__":
    # Create sample application
    app = LoanApplication(
        units=8,
        address="307 S Main Street",
        city="Hopkinsville",
        state="KY",
        zip_code="44240",
        estimated_value=1200000,
        purchase_price=980000,
        loan_amount=784000,
        note_type="30 YR Fixed",
        points_to_lender=1.0,
        credit_score_1=688,
        credit_score_2=712,
        credit_score_3=703,
        property_type="Multifamily",
        asset_class="C"
    )
    
    # Process through sizer
    processor = SizerProcessor("../demo-data/template.xlsx")
    result = processor.process_application(app, interest_rate=8.50)
    
    print(f"Decision: {result.overall_decision}")
    print(f"Reason: {result.decision_reason}")
    print(f"Processing time: {result.processing_time:.2f} seconds")
    print(f"\nPrograms:")
    for prog in result.programs:
        print(f"  {prog.name}: {prog.status}")
