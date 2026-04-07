"""
Loan Sizer Automation System
Backend API for processing loan applications from emails
"""

import os
import json
import re
import asyncio
from typing import List, Optional, Dict, Any
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field, validator
import anthropic
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

app = FastAPI(title="Loan Sizer Automation API", version="1.0.0")

# CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Claude client
claude_client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY", ""))

# ==================== DATA MODELS ====================

class CreditScores(BaseModel):
    score1: int = Field(..., ge=300, le=850)
    score2: int = Field(..., ge=300, le=850)
    score3: int = Field(..., ge=300, le=850)
    
    @property
    def middle_score(self) -> int:
        """Return the middle credit score (standard lending practice)"""
        scores = sorted([self.score1, self.score2, self.score3])
        return scores[1]

class LoanApplication(BaseModel):
    # Property Details
    units: int = Field(..., ge=1, le=100)
    unit_size: Optional[str] = None
    address: str
    city: str
    state: str
    zip_code: str
    
    # Financial Details
    estimated_value: float = Field(..., gt=0)
    purchase_price: float = Field(..., gt=0)
    loan_amount: float = Field(..., gt=0)
    note_type: str  # e.g., "30 YR Fixed", "15 YR Fixed"
    points_to_lender: float = Field(default=0.0, ge=0, le=10)
    
    # Credit
    credit_scores: CreditScores
    
    # Rate (fetched daily)
    daily_rate: Optional[float] = None
    
    @property
    def ltv_ratio(self) -> float:
        """Loan-to-Value ratio"""
        return (self.loan_amount / self.estimated_value) * 100
    
    @property
    def full_address(self) -> str:
        return f"{self.address}, {self.city}, {self.state} {self.zip_code}"

class ExtractionResult(BaseModel):
    application: LoanApplication
    raw_extraction: Dict[str, Any]
    confidence: float
    missing_fields: List[str]

class ProgramResult(BaseModel):
    program_name: str
    status: str  # "PASS" or "FAIL"
    reason: Optional[str] = None
    max_loan_amount: Optional[float] = None
    dscr: Optional[float] = None

class SizerResult(BaseModel):
    application: LoanApplication
    programs: List[ProgramResult]
    overall_decision: str  # "APPROVE", "REJECT", or "REVIEW"
    decision_reason: str
    processed_at: datetime
    
class EmailDraft(BaseModel):
    subject: str
    body: str
    recipient: str
    approval_status: str
    key_points: List[str]

# ==================== EMAIL EXTRACTION SERVICE ====================

EXTRACTION_PROMPT = """You are an expert loan application data extractor. Your job is to parse emails from loan applicants and extract structured data.

Extract the following fields from the email below:

REQUIRED FIELDS:
- units: Number of units (integer)
- unit_size: Size of units (e.g., "750 sq ft", "1200 sq ft each") - optional
- address: Street address
- city: City name
- state: State abbreviation or full name
- zip_code: ZIP code
- estimated_value: Estimated property value (number, no $ or commas)
- purchase_price: Purchase price (number, no $ or commas)
- loan_amount: Requested loan amount (number, no $ or commas)
- note_type: Type of note (e.g., "30 YR Fixed", "15 YR Fixed", "ARM")
- points_to_lender: Points percentage (number, default 0 if not mentioned)
- credit_score_1: First credit score (3-digit number)
- credit_score_2: Second credit score (3-digit number)
- credit_score_3: Third credit score (3-digit number)

EMAIL CONTENT:
{email_content}

Respond ONLY with a JSON object in this exact format:
{{
    "units": 8,
    "unit_size": "750 sq ft each",
    "address": "1428 Elmwood Ave",
    "city": "Philadelphia",
    "state": "PA",
    "zip_code": "19103",
    "estimated_value": 1200000,
    "purchase_price": 980000,
    "loan_amount": 784000,
    "note_type": "30 YR Fixed",
    "points_to_lender": 1.0,
    "credit_score_1": 688,
    "credit_score_2": 712,
    "credit_score_3": 703
}}

If a field is missing, use null. Be precise with numbers - remove $ signs and commas."""

async def extract_from_email(email_content: str) -> ExtractionResult:
    """Use Claude to extract loan application data from email"""
    
    if not os.getenv("ANTHROPIC_API_KEY"):
        # Demo mode - use regex extraction
        return await extract_with_regex(email_content)
    
    try:
        response = claude_client.messages.create(
            model="claude-3-opus-20240229",
            max_tokens=1000,
            temperature=0,
            messages=[{
                "role": "user",
                "content": EXTRACTION_PROMPT.format(email_content=email_content)
            }]
        )
        
        # Parse JSON response
        extracted = json.loads(response.content[0].text)
        
        # Check for missing required fields
        required_fields = ['units', 'address', 'city', 'state', 'zip_code', 
                          'estimated_value', 'purchase_price', 'loan_amount',
                          'note_type', 'credit_score_1', 'credit_score_2', 'credit_score_3']
        missing = [f for f in required_fields if extracted.get(f) is None]
        
        # Build application object
        application = LoanApplication(
            units=extracted['units'],
            unit_size=extracted.get('unit_size'),
            address=extracted['address'],
            city=extracted['city'],
            state=extracted['state'],
            zip_code=extracted['zip_code'],
            estimated_value=float(extracted['estimated_value']),
            purchase_price=float(extracted['purchase_price']),
            loan_amount=float(extracted['loan_amount']),
            note_type=extracted['note_type'],
            points_to_lender=float(extracted.get('points_to_lender', 0)),
            credit_scores=CreditScores(
                score1=int(extracted['credit_score_1']),
                score2=int(extracted['credit_score_2']),
                score3=int(extracted['credit_score_3'])
            )
        )
        
        return ExtractionResult(
            application=application,
            raw_extraction=extracted,
            confidence=0.95 if not missing else 0.75,
            missing_fields=missing
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")

async def extract_with_regex(email_content: str) -> ExtractionResult:
    """Fallback extraction using regex patterns"""
    
    text = email_content
    
    # Extract patterns
    patterns = {
        'units': r'(?i)(\d+)\s*units?',
        'unit_size': r'(?i)(\d+)\s*sq\s*ft(?:\s*each)?',
        'address': r'(?i)(\d+\s+[^,]+(?:Ave|St|Dr|Rd|Blvd|Ln|Way|Ct))',
        'city': r'(?i)(?:city[:\s]+)?([A-Za-z\s]+)(?=,\s*[A-Z]{2})',
        'state': r'(?i),\s*([A-Z]{2})\s+\d{5}',
        'zip': r'(?i)\b(\d{5}(?:-\d{4})?)\b',
        'estimated_value': r'(?i)estimated value[:\s]+\$?([\d,]+)',
        'purchase_price': r'(?i)purchase price[:\s]+\$?([\d,]+)',
        'loan_amount': r'(?i)loan amount[:\s]+\$?([\d,]+)',
        'note_type': r'(?i)(30|15|20)\s*YR\s*(?:Fixed|ARM)',
        'points': r'(?i)points[:\s]+(\d+(?:\.\d+)?)',
        'credit_scores': r'(?i)(\d{3})[\s,]+(\d{3})[\s,]+(\d{3})'
    }
    
    extracted = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            if key == 'credit_scores':
                extracted['credit_score_1'] = int(match.group(1))
                extracted['credit_score_2'] = int(match.group(2))
                extracted['credit_score_3'] = int(match.group(3))
            elif key in ['estimated_value', 'purchase_price', 'loan_amount']:
                extracted[key] = float(match.group(1).replace(',', ''))
            elif key == 'units':
                extracted[key] = int(match.group(1))
            elif key == 'points':
                extracted[key] = float(match.group(1))
            elif key == 'note_type':
                extracted[key] = f"{match.group(1)} YR Fixed"
            else:
                extracted[key] = match.group(1).strip()
    
    # Build application
    missing = []
    required = ['units', 'address', 'city', 'state', 'zip', 
                'estimated_value', 'purchase_price', 'loan_amount',
                'note_type', 'credit_score_1', 'credit_score_2', 'credit_score_3']
    
    for field in required:
        if field not in extracted:
            missing.append(field)
    
    if missing:
        raise HTTPException(status_code=422, detail=f"Missing required fields: {missing}")
    
    application = LoanApplication(
        units=extracted['units'],
        unit_size=extracted.get('unit_size'),
        address=extracted['address'],
        city=extracted.get('city', 'Unknown'),
        state=extracted.get('state', 'XX'),
        zip_code=extracted.get('zip', '00000'),
        estimated_value=extracted['estimated_value'],
        purchase_price=extracted['purchase_price'],
        loan_amount=extracted['loan_amount'],
        note_type=extracted['note_type'],
        points_to_lender=extracted.get('points', 0),
        credit_scores=CreditScores(
            score1=extracted['credit_score_1'],
            score2=extracted['credit_score_2'],
            score3=extracted['credit_score_3']
        )
    )
    
    return ExtractionResult(
        application=application,
        raw_extraction=extracted,
        confidence=0.85,
        missing_fields=[]
    )

# ==================== EXCEL SIZER SERVICE ====================

class SizerService:
    """Service for populating and reading the Excel SIZER tab"""
    
    def __init__(self, template_path: str):
        self.template_path = Path(template_path)
        
    def populate_sizer(self, application: LoanApplication, daily_rate: float, output_path: str) -> SizerResult:
        """Populate SIZER tab with application data and return results"""
        
        # Load workbook
        wb = load_workbook(self.template_path, data_only=False)
        
        # Get SIZER tab (or first sheet if not found)
        sizer_sheet = wb['SIZER'] if 'SIZER' in wb.sheetnames else wb.active
        
        # Map application fields to cell coordinates
        # These mappings would be customized based on the actual Excel file
        cell_mappings = {
            'C8': application.units,
            'C9': application.unit_size or '',
            'E5': application.address,
            'E6': application.city,
            'E7': application.state,
            'E8': application.zip_code,
            'G5': application.estimated_value,
            'G6': application.purchase_price,
            'I5': application.loan_amount,
            'I8': application.note_type,
            'I10': application.points_to_lender,
            'M7': application.credit_scores.middle_score,
            'G10': daily_rate,  # Daily rate
        }
        
        # Write values to cells
        for cell, value in cell_mappings.items():
            sizer_sheet[cell] = value
        
        # Save workbook
        wb.save(output_path)
        
        # In production, you would:
        # 1. Use LibreOffice headless to recalculate formulas
        # 2. Read back the Programs section results
        # For now, simulate the evaluation
        
        programs = self._evaluate_programs(application, daily_rate)
        
        # Determine overall decision
        pass_count = sum(1 for p in programs if p.status == "PASS")
        if pass_count >= 2:
            decision = "APPROVE"
            reason = f"Qualified for {pass_count} of 3 programs"
        elif pass_count == 1:
            decision = "REVIEW"
            reason = "Marginal qualification - manual review recommended"
        else:
            decision = "REJECT"
            reason = "Does not meet minimum program requirements"
        
        return SizerResult(
            application=application,
            programs=programs,
            overall_decision=decision,
            decision_reason=reason,
            processed_at=datetime.now()
        )
    
    def _evaluate_programs(self, application: LoanApplication, rate: float) -> List[ProgramResult]:
        """Evaluate application against program guidelines"""
        
        programs = []
        
        # Program 1: Short Term Note Sale
        if application.ltv_ratio <= 75 and application.credit_scores.middle_score >= 680:
            programs.append(ProgramResult(
                program_name="Short Term Note Sale",
                status="PASS",
                max_loan_amount=application.estimated_value * 0.75,
                dscr=1.25
            ))
        else:
            fail_reasons = []
            if application.ltv_ratio > 75:
                fail_reasons.append(f"LTV {application.ltv_ratio:.1f}% exceeds 75% max")
            if application.credit_scores.middle_score < 680:
                fail_reasons.append(f"Credit score {application.credit_scores.middle_score} below 680 minimum")
            
            programs.append(ProgramResult(
                program_name="Short Term Note Sale",
                status="FAIL",
                reason="; ".join(fail_reasons)
            ))
        
        # Program 2: Insurance Program
        if application.ltv_ratio <= 80 and application.credit_scores.middle_score >= 640:
            programs.append(ProgramResult(
                program_name="Insurance Program",
                status="PASS",
                max_loan_amount=application.estimated_value * 0.80,
                dscr=1.20
            ))
        else:
            programs.append(ProgramResult(
                program_name="Insurance Program",
                status="FAIL",
                reason="LTV or credit score below requirements"
            ))
        
        # Program 3: Long Term Note Sale
        if application.ltv_ratio <= 70 and application.credit_scores.middle_score >= 720:
            programs.append(ProgramResult(
                program_name="Long Term Note Sale",
                status="PASS",
                max_loan_amount=application.estimated_value * 0.70,
                dscr=1.30
            ))
        else:
            programs.append(ProgramResult(
                program_name="Long Term Note Sale",
                status="FAIL",
                reason="LTV or credit score below requirements"
            ))
        
        return programs

# ==================== EMAIL GENERATION SERVICE ====================

def generate_approval_email(result: SizerResult, applicant_email: str) -> EmailDraft:
    """Generate approval email with loan terms"""
    
    passed_programs = [p for p in result.programs if p.status == "PASS"]
    best_program = passed_programs[0] if passed_programs else None
    
    subject = f"Loan Application Approved - {result.application.full_address}"
    
    body = f"""Dear Applicant,

Congratulations! Your loan application for the property at {result.application.full_address} has been APPROVED.

LOAN DETAILS:
• Property: {result.application.units} units, {result.application.unit_size or 'N/A'}
• Loan Amount: ${result.application.loan_amount:,.2f}
• Estimated Value: ${result.application.estimated_value:,.2f}
• LTV Ratio: {result.application.ltv_ratio:.2f}%
• Credit Score Used: {result.application.credit_scores.middle_score} (middle of 3)

APPROVED PROGRAMS:
"""
    
    for program in passed_programs:
        body += f"• {program.program_name}\n"
        if program.max_loan_amount:
            body += f"  - Max Loan Amount: ${program.max_loan_amount:,.2f}\n"
        if program.dscr:
            body += f"  - DSCR: {program.dscr:.2f}\n"
    
    body += f"""
NEXT STEPS:
1. Review and sign the attached loan commitment letter
2. Submit required documentation (income verification, property appraisal)
3. Schedule closing with our processing team

Please reply to this email or call us at (555) 123-4567 to proceed.

Best regards,
Loan Processing Team
"""
    
    return EmailDraft(
        subject=subject,
        body=body,
        recipient=applicant_email,
        approval_status="APPROVED",
        key_points=[f"Approved for {len(passed_programs)} programs", f"LTV: {result.application.ltv_ratio:.1f}%"]
    )

def generate_rejection_email(result: SizerResult, applicant_email: str) -> EmailDraft:
    """Generate rejection email with explanation"""
    
    failed_programs = [p for p in result.programs if p.status == "FAIL"]
    
    subject = f"Loan Application Update - {result.application.full_address}"
    
    body = f"""Dear Applicant,

Thank you for your loan application for the property at {result.application.full_address}.

After careful review, we regret to inform you that your application does not meet our current program requirements at this time.

APPLICATION DETAILS:
• Property: {result.application.units} units
• Requested Loan: ${result.application.loan_amount:,.2f}
• Estimated Value: ${result.application.estimated_value:,.2f}
• LTV Ratio: {result.application.ltv_ratio:.2f}%
• Credit Score: {result.application.credit_scores.middle_score}

REASONS:
"""
    
    for program in failed_programs:
        if program.reason:
            body += f"• {program.program_name}: {program.reason}\n"
    
    body += f"""
RECOMMENDATIONS:
You may reapply if:
• You can increase your down payment to lower the LTV ratio
• Your credit score improves
• You provide additional collateral

We appreciate your interest and wish you success with your investment.

Best regards,
Loan Processing Team
"""
    
    return EmailDraft(
        subject=subject,
        body=body,
        recipient=applicant_email,
        approval_status="DECLINED",
        key_points=["Does not meet program requirements", result.decision_reason]
    )

# ==================== API ENDPOINTS ====================

@app.get("/")
async def root():
    return {"status": "Loan Sizer Automation API", "version": "1.0.0"}

@app.post("/api/extract", response_model=ExtractionResult)
async def extract_application(email_content: str = Form(...)):
    """Extract loan application data from email content"""
    return await extract_from_email(email_content)

@app.post("/api/process")
async def process_application(
    background_tasks: BackgroundTasks,
    email_content: str = Form(...),
    daily_rate: float = Form(...),
    applicant_email: str = Form(...),
    template_path: Optional[str] = Form(None)
):
    """Full pipeline: extract → populate sizer → evaluate → generate email"""
    
    # Step 1: Extract data from email
    extraction = await extract_from_email(email_content)
    
    if extraction.missing_fields:
        return {
            "status": "incomplete",
            "message": f"Missing required fields: {extraction.missing_fields}",
            "extraction": extraction
        }
    
    # Step 2: Populate SIZER and evaluate
    template = template_path or "template.xlsx"
    sizer_service = SizerService(template)
    
    output_path = f"/tmp/sizer_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    result = sizer_service.populate_sizer(extraction.application, daily_rate, output_path)
    
    # Step 3: Generate email
    if result.overall_decision == "APPROVE":
        email = generate_approval_email(result, applicant_email)
    elif result.overall_decision == "REJECT":
        email = generate_rejection_email(result, applicant_email)
    else:
        # REVIEW - generate neutral email
        email = EmailDraft(
            subject=f"Loan Application Under Review - {result.application.full_address}",
            body="Your application is under manual review...",
            recipient=applicant_email,
            approval_status="UNDER REVIEW",
            key_points=["Manual review required"]
        )
    
    return {
        "status": "complete",
        "extraction": extraction,
        "sizer_result": result,
        "generated_email": email,
        "output_file": output_path,
        "time_saved": "20-25 minutes"
    }

@app.post("/api/daily-rate")
async def update_daily_rate(rate: float, effective_date: Optional[str] = None):
    """Update daily interest rate"""
    # In production, this would store to database
    return {
        "status": "updated",
        "daily_rate": rate,
        "effective_date": effective_date or datetime.now().isoformat()
    }

@app.get("/api/demo/scenarios")
async def get_demo_scenarios():
    """Return demo application scenarios for testing"""
    return {
        "scenarios": [
            {
                "name": "APPROVE - Strong Application",
                "email": """Subject: Loan Application – 1428 Elmwood Ave

Hi,

I'm interested in a loan for a multifamily property at 1428 Elmwood Ave, Philadelphia, PA 19103.
Property details:
- 8 units, approx 750 sq ft each
- Estimated value: $1,200,000
- Purchase price: $980,000
- Loan amount requested: $784,000
- Note type: 30 YR Fixed

My three credit scores are 688, 712, and 703.
Points to lender: 1

Thanks,
James Whitfield""",
                "expected_outcome": "APPROVE",
                "rate": 8.50
            },
            {
                "name": "REJECT - High LTV",
                "email": """Subject: Loan Application Request

Hello,

Looking for financing on:
- 12 units at 456 Oak Street, Chicago, IL 60601
- Value: $800,000
- Purchase: $750,000  
- Loan needed: $700,000
- Note: 30 YR Fixed
- Credit scores: 620, 641, 655

Thanks,
Michael Chen""",
                "expected_outcome": "REJECT",
                "rate": 8.75
            },
            {
                "name": "BORDERLINE - Review Required",
                "email": """Subject: Quick loan app

Property: 789 Pine Ave, Miami FL 33101
4 units, 900 sq ft each
Value: $600k
Buy: $520k
Loan: $450k
Scores: 665, 678, 682
30 YR Fixed

Thanks""",
                "expected_outcome": "REVIEW",
                "rate": 8.60
            }
        ]
    }

# ==================== MAIN ====================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=5050)
