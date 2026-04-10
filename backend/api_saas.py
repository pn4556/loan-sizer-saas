"""
Loan Sizer SaaS API
Multi-tenant API with authentication, file uploads, and PDF parsing
"""

import os
import json
import shutil
import hashlib
from datetime import datetime, timedelta
from typing import List, Optional
from pathlib import Path

from fastapi import FastAPI, Depends, HTTPException, status, File, UploadFile, Form, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel

# Import our modules
from database import get_db, init_database
from models import (
    Client, User, ExcelTemplate, LoanApplication, 
    ApiKey, AuditLog
)
from auth import (
    authenticate_user, create_access_token, create_refresh_token,
    get_current_user, require_admin, register_client,
    get_password_hash, verify_password
)
from processor_custom import SizerProcessor, LoanApplication as LoanAppModel
from pdf_parser import PDFLoanParser
from file_parser import UniversalFileParser, parse_loan_file

# Import email processing router
from email_api import router as email_router

# Initialize file parser
file_parser = UniversalFileParser()

# Initialize app
app = FastAPI(
    title="Loan Sizer SaaS Platform",
    version="2.1.0",
    description="Multi-tenant loan processing with universal file support and Excel export"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Setup directories
UPLOAD_DIR = Path("./uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
TEMPLATES_DIR = UPLOAD_DIR / "templates"
TEMPLATES_DIR.mkdir(exist_ok=True)
PDFS_DIR = UPLOAD_DIR / "pdfs"
PDFS_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR = UPLOAD_DIR / "outputs"
OUTPUTS_DIR.mkdir(exist_ok=True)

# Include email processing router
app.include_router(email_router)

# Startup event
@app.on_event("startup")
async def startup_event():
    print("🚀 Loan Sizer API starting up...")
    print(f"📍 Root path: /")
    print(f"📍 Health path: /health")
    print(f"📍 API docs: /docs")
    
    # Initialize database tables
    try:
        init_database()
        print("✅ Database initialized")
    except Exception as e:
        print(f"⚠️ Database initialization warning: {e}")

# ==================== ROOT ROUTE ====================

@app.get("/")
async def root():
    """Root endpoint - API info"""
    return {
        "service": "Loan Sizer SaaS Platform",
        "version": "2.0.0",
        "status": "running",
        "docs": "/docs",
        "health": "/health"
    }

@app.get("/health")
async def health():
    """Health check endpoint"""
    return {"status": "healthy", "service": "loan-sizer-api"}

# ==================== AUTH SCHEMAS ====================

class LoginRequest(BaseModel):
    email: str
    password: str

class LoginResponse(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str = "bearer"
    user: dict

class RegisterRequest(BaseModel):
    company_name: str
    email: str
    password: str
    first_name: str
    last_name: str

class CreateUserRequest(BaseModel):
    email: str
    password: str
    first_name: str
    last_name: str
    role: str = "loan_officer"

# ==================== AUTH ROUTES ====================

@app.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest, db: Session = Depends(get_db)):
    """Authenticate user and return tokens"""
    user = authenticate_user(db, request.email, request.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid email or password"
        )
    
    # Update last login
    user.last_login = datetime.utcnow()
    db.commit()
    
    # Create tokens
    access_token = create_access_token(
        data={"sub": str(user.id), "client_id": user.client_id, "role": user.role}
    )
    refresh_token = create_refresh_token(
        data={"sub": str(user.id), "client_id": user.client_id}
    )
    
    return LoginResponse(
        access_token=access_token,
        refresh_token=refresh_token,
        user={
            "id": user.id,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "role": user.role,
            "client": {
                "id": user.client.id,
                "company_name": user.client.company_name,
                "slug": user.client.slug
            }
        }
    )

@app.post("/auth/register")
async def register(request: RegisterRequest, db: Session = Depends(get_db)):
    """Register a new client"""
    # Check if email exists
    existing = db.query(User).filter(User.email == request.email).first()
    if existing:
        raise HTTPException(400, "Email already registered")
    
    # Create client and admin user
    client, admin = register_client(
        db=db,
        company_name=request.company_name,
        email=request.email,
        password=request.password,
        admin_first_name=request.first_name,
        admin_last_name=request.last_name
    )
    
    return {
        "message": "Registration successful",
        "client_id": client.id,
        "client_slug": client.slug,
        "trial_ends": client.trial_ends_at
    }

@app.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    """Get current user info"""
    return {
        "id": current_user.id,
        "email": current_user.email,
        "first_name": current_user.first_name,
        "last_name": current_user.last_name,
        "role": current_user.role,
        "client": {
            "id": current_user.client.id,
            "company_name": current_user.client.company_name,
            "slug": current_user.client.slug,
            "plan": current_user.client.plan,
            "trial_ends": current_user.client.trial_ends_at
        }
    }

# ==================== USER MANAGEMENT (Admin) ====================

@app.post("/admin/users")
async def create_user(
    request: CreateUserRequest,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Create a new user in the organization (Admin only)"""
    
    # Check email uniqueness within client
    existing = db.query(User).filter(
        User.email == request.email,
        User.client_id == current_user.client_id
    ).first()
    if existing:
        raise HTTPException(400, "Email already exists in your organization")
    
    user = User(
        client_id=current_user.client_id,
        email=request.email,
        first_name=request.first_name,
        last_name=request.last_name,
        role=request.role
    )
    user.set_password(request.password)
    
    db.add(user)
    db.commit()
    
    return {"id": user.id, "email": user.email, "message": "User created"}

@app.get("/admin/users")
async def list_users(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """List all users in organization (Admin only)"""
    users = db.query(User).filter(User.client_id == current_user.client_id).all()
    return [{
        "id": u.id,
        "email": u.email,
        "first_name": u.first_name,
        "last_name": u.last_name,
        "role": u.role,
        "is_active": u.is_active,
        "last_login": u.last_login
    } for u in users]

# ==================== TEMPLATE MANAGEMENT ====================

@app.post("/templates/upload")
async def upload_template(
    name: str = Form(...),
    description: str = Form(None),
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Upload Excel template (Admin only)"""
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
        raise HTTPException(400, "Only Excel files (.xlsx, .xlsm, .xls) allowed")
    
    # Read file content
    content = await file.read()
    
    # Calculate hash for deduplication
    file_hash = hashlib.sha256(content).hexdigest()
    
    # Save file
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"{current_user.client_id}_{timestamp}_{file.filename}"
    file_path = TEMPLATES_DIR / filename
    
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Analyze template (extract cell mappings)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=False)
        
        # Find input cells (simplified analysis)
        cell_mappings = {}
        if 'SIZER' in wb.sheetnames:
            sheet = wb['SIZER']
            # This would be expanded with full analysis
            cell_mappings = {
                "C8": "units",
                "E5": "address",
                "E6": "city",
                "E7": "state",
                "E8": "zip_code",
                "G5": "estimated_value",
                "G6": "purchase_price",
                "J4": "loan_amount",
                "I8": "note_type",
                "I10": "points_to_lender",
                "M7": "credit_score",
            }
        wb.close()
    except Exception as e:
        # Delete file if analysis fails
        file_path.unlink(missing_ok=True)
        raise HTTPException(400, f"Could not analyze Excel file: {str(e)}")
    
    # Create template record
    template = ExcelTemplate(
        client_id=current_user.client_id,
        name=name,
        description=description,
        file_path=str(file_path),
        file_size=len(content),
        file_hash=file_hash,
        cell_mappings=cell_mappings
    )
    
    # If first template, make it default
    existing = db.query(ExcelTemplate).filter(
        ExcelTemplate.client_id == current_user.client_id
    ).first()
    if not existing:
        template.is_default = True
    
    db.add(template)
    db.commit()
    
    return {
        "id": template.id,
        "name": template.name,
        "cell_mappings": template.cell_mappings,
        "message": "Template uploaded successfully"
    }

@app.get("/templates")
async def list_templates(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """List all templates for the organization"""
    templates = db.query(ExcelTemplate).filter(
        ExcelTemplate.client_id == current_user.client_id
    ).all()
    
    return [{
        "id": t.id,
        "name": t.name,
        "description": t.description,
        "is_default": t.is_default,
        "is_active": t.is_active,
        "file_size": t.file_size,
        "created_at": t.created_at,
        "cell_count": len(t.cell_mappings) if t.cell_mappings else 0
    } for t in templates]

@app.post("/templates/{template_id}/set-default")
async def set_default_template(
    template_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Set template as default (Admin only)"""
    
    # Unset current default
    db.query(ExcelTemplate).filter(
        ExcelTemplate.client_id == current_user.client_id
    ).update({"is_default": False})
    
    # Set new default
    template = db.query(ExcelTemplate).filter(
        ExcelTemplate.id == template_id,
        ExcelTemplate.client_id == current_user.client_id
    ).first()
    
    if not template:
        raise HTTPException(404, "Template not found")
    
    template.is_default = True
    db.commit()
    
    return {"message": "Default template updated"}

# ==================== APPLICATION PROCESSING ====================

class ProcessFromEmailRequest(BaseModel):
    email_content: str
    interest_rate: float
    applicant_email: str
    applicant_name: Optional[str] = None
    template_id: Optional[int] = None
    extracted_data: Optional[dict] = None  # Allow frontend to provide parsed data

@app.post("/applications/process/email")
async def process_from_email(
    request: ProcessFromEmailRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Process loan application from email text"""
    
    # Get template
    if request.template_id:
        template = db.query(ExcelTemplate).filter(
            ExcelTemplate.id == request.template_id,
            ExcelTemplate.client_id == current_user.client_id
        ).first()
    else:
        template = db.query(ExcelTemplate).filter(
            ExcelTemplate.client_id == current_user.client_id,
            ExcelTemplate.is_default == True
        ).first()
    
    if not template:
        raise HTTPException(400, "No template available. Please upload an Excel template first.")
    
    # Create application record
    app_record = LoanApplication(
        client_id=current_user.client_id,
        template_id=template.id,
        processed_by_id=current_user.id,
        source_type="email",
        source_email=request.email_content,
        applicant_email=request.applicant_email,
        interest_rate=request.interest_rate,
        status="processing"
    )
    db.add(app_record)
    db.commit()
    
    try:
        # Use provided extracted data if available, else extract
        if request.extracted_data:
            extracted = request.extracted_data
            app_record.extraction_method = "frontend"
        else:
            # Extract data (use AI if available, else regex)
            import anthropic
            try:
                client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
                # AI extraction would go here
                extracted = _extract_with_regex(request.email_content)  # Fallback for now
                app_record.extraction_method = "regex"
            except:
                extracted = _extract_with_regex(request.email_content)
                app_record.extraction_method = "regex"
        
        # Validate extraction
        required = ['units', 'address', 'city', 'state', 'zip_code',
                   'estimated_value', 'purchase_price', 'loan_amount',
                   'credit_score_1', 'credit_score_2', 'credit_score_3']
        missing = [f for f in required if f not in extracted]
        app_record.missing_fields = missing
        
        if missing:
            app_record.status = "review"
            app_record.officer_notes = f"Missing fields: {', '.join(missing)}"
            db.commit()
            return {
                "application_id": app_record.id,
                "status": "review",
                "message": f"Missing required fields: {', '.join(missing)}",
                "extracted_data": extracted
            }
        
        # Create loan application model
        loan_app = LoanAppModel(**extracted)
        
        # Process through sizer
        processor = SizerProcessor(template.file_path)
        result = processor.process_application(loan_app, request.interest_rate)
        
        # Update record with results
        app_record.units = loan_app.units
        app_record.property_address = loan_app.address
        app_record.property_city = loan_app.city
        app_record.property_state = loan_app.state
        app_record.property_zip = loan_app.zip_code
        app_record.estimated_value = loan_app.estimated_value
        app_record.purchase_price = loan_app.purchase_price
        app_record.loan_amount = loan_app.loan_amount
        app_record.note_type = loan_app.note_type
        app_record.points_to_lender = loan_app.points_to_lender
        app_record.credit_score_1 = loan_app.credit_score_1
        app_record.credit_score_2 = loan_app.credit_score_2
        app_record.credit_score_3 = loan_app.credit_score_3
        app_record.credit_score_middle = loan_app.credit_score_middle
        app_record.ltv_ratio = loan_app.ltv_ratio
        app_record.extraction_confidence = 0.85
        app_record.programs_results = [{
            "program_name": p.program_name,
            "status": p.status,
            "max_loan": p.max_loan_amount,
            "dscr": p.dscr,
            "reason": p.reason
        } for p in result.programs]
        app_record.overall_decision = result.overall_decision
        app_record.decision_reason = result.decision_reason
        app_record.output_excel_path = result.output_file
        app_record.processing_time_seconds = result.processing_time
        app_record.status = "review"  # Awaiting officer approval
        
        db.commit()
        
        return {
            "application_id": app_record.id,
            "status": "success",
            "decision": result.overall_decision,
            "decision_reason": result.decision_reason,
            "processing_time": result.processing_time,
            "programs": app_record.programs_results,
            "output_file": result.output_file,
            "application": {
                "units": loan_app.units,
                "address": loan_app.full_address,
                "loan_amount": loan_app.loan_amount,
                "ltv_ratio": loan_app.ltv_ratio,
                "credit_score": loan_app.credit_score_middle
            }
        }
        
    except Exception as e:
        app_record.status = "error"
        app_record.officer_notes = str(e)
        db.commit()
        raise HTTPException(500, f"Processing failed: {str(e)}")


@app.post("/applications/process/pdf")
async def process_from_pdf(
    interest_rate: float = Form(...),
    applicant_email: str = Form(...),
    file: UploadFile = File(...),
    template_id: Optional[int] = Form(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Process loan application from PDF file"""
    
    # Validate file
    if not file.content_type == "application/pdf":
        raise HTTPException(400, "Only PDF files allowed")
    
    # Get template
    if template_id:
        template = db.query(ExcelTemplate).filter(
            ExcelTemplate.id == template_id,
            ExcelTemplate.client_id == current_user.client_id
        ).first()
    else:
        template = db.query(ExcelTemplate).filter(
            ExcelTemplate.client_id == current_user.client_id,
            ExcelTemplate.is_default == True
        ).first()
    
    if not template:
        raise HTTPException(400, "No template available")
    
    # Save PDF
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    pdf_filename = f"{current_user.client_id}_{timestamp}_{file.filename}"
    pdf_path = PDFS_DIR / pdf_filename
    
    content = await file.read()
    with open(pdf_path, "wb") as f:
        f.write(content)
    
    # Create application record
    app_record = LoanApplication(
        client_id=current_user.client_id,
        template_id=template.id,
        processed_by_id=current_user.id,
        source_type="pdf",
        source_pdf_path=str(pdf_path),
        applicant_email=applicant_email,
        interest_rate=interest_rate,
        status="processing"
    )
    db.add(app_record)
    db.commit()
    
    try:
        # Parse PDF
        parser = PDFLoanParser()
        pdf_result = parser.parse_pdf(content)
        
        if not pdf_result.success:
            app_record.status = "review"
            app_record.officer_notes = f"PDF parsing failed: {pdf_result.error_message}"
            db.commit()
            raise HTTPException(400, f"Could not parse PDF: {pdf_result.error_message}")
        
        # Convert extracted fields to loan application
        extracted = pdf_result.fields
        extracted['points_to_lender'] = extracted.get('points', 0.0)
        
        # Create and process
        loan_app = LoanAppModel(**extracted)
        processor = SizerProcessor(template.file_path)
        result = processor.process_application(loan_app, interest_rate)
        
        # Update record
        app_record.extraction_method = "pdf"
        app_record.extraction_confidence = pdf_result.confidence
        app_record.status = "review"
        db.commit()
        
        return {
            "application_id": app_record.id,
            "status": "success",
            "pdf_confidence": pdf_result.confidence,
            "decision": result.overall_decision,
            "extracted_fields": list(pdf_result.fields.keys())
        }
        
    except Exception as e:
        app_record.status = "error"
        app_record.officer_notes = str(e)
        db.commit()
        raise HTTPException(500, f"Processing failed: {str(e)}")


# ==================== DASHBOARD ROUTES ====================

@app.get("/dashboard/stats")
async def get_dashboard_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get real dashboard statistics for the organization"""
    
    client_id = current_user.client_id
    
    # Total applications
    total_apps = db.query(LoanApplication).filter(
        LoanApplication.client_id == client_id
    ).count()
    
    # This month's applications
    from datetime import datetime, timedelta
    month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_apps = db.query(LoanApplication).filter(
        LoanApplication.client_id == client_id,
        LoanApplication.created_at >= month_start
    ).count()
    
    # Last month's applications (for comparison)
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)
    last_month_apps = db.query(LoanApplication).filter(
        LoanApplication.client_id == client_id,
        LoanApplication.created_at >= last_month_start,
        LoanApplication.created_at < month_start
    ).count()
    
    # Decision breakdown
    decisions = db.query(LoanApplication.overall_decision, 
                        func.count(LoanApplication.id)).filter(
        LoanApplication.client_id == client_id
    ).group_by(LoanApplication.overall_decision).all()
    
    decision_counts = {d[0] or "pending": d[1] for d in decisions}
    approved_count = decision_counts.get("APPROVE", 0) + decision_counts.get("approve", 0)
    rejected_count = decision_counts.get("REJECT", 0) + decision_counts.get("reject", 0)
    pending_count = decision_counts.get("REVIEW", 0) + decision_counts.get("review", 0) + decision_counts.get("pending", 0)
    
    # Average processing time
    avg_time = db.query(func.avg(LoanApplication.processing_time_seconds)).filter(
        LoanApplication.client_id == client_id,
        LoanApplication.processing_time_seconds != None
    ).scalar()
    
    # Time saved calculation (23 min saved per app vs manual)
    time_saved_minutes = total_apps * 23
    time_saved_hours = round(time_saved_minutes / 60, 1)
    
    # Calculate month-over-month change
    month_change = 0
    if last_month_apps > 0:
        month_change = round(((month_apps - last_month_apps) / last_month_apps) * 100)
    elif month_apps > 0:
        month_change = 100
    
    return {
        "total_applications": total_apps,
        "this_month": month_apps,
        "last_month": last_month_apps,
        "month_change_percent": month_change,
        "decisions": decision_counts,
        "approved": approved_count,
        "rejected": rejected_count,
        "pending_review": pending_count,
        "average_processing_time": round(avg_time, 2) if avg_time else 0,
        "time_saved_minutes": time_saved_minutes,
        "time_saved_hours": time_saved_hours,
        "efficiency_gain": 93  # Percentage
    }


@app.get("/dashboard/activity")
async def get_recent_activity(
    limit: int = 10,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get recent activity for the dashboard"""
    
    client_id = current_user.client_id
    
    apps = db.query(LoanApplication).filter(
        LoanApplication.client_id == client_id
    ).order_by(LoanApplication.created_at.desc()).limit(limit).all()
    
    activities = []
    for app in apps:
        # Determine icon and status based on decision
        if app.overall_decision == "APPROVE":
            icon = "success"
            title = "Application Approved"
        elif app.overall_decision == "REJECT":
            icon = "error"
            title = "Application Rejected"
        elif app.status == "processing":
            icon = "pending"
            title = "Processing Started"
        else:
            icon = "pending"
            title = "Under Review"
        
        # Format time ago
        time_diff = datetime.utcnow() - app.created_at
        if time_diff.days > 0:
            time_ago = f"{time_diff.days} day{'s' if time_diff.days > 1 else ''} ago"
        elif time_diff.seconds > 3600:
            hours = time_diff.seconds // 3600
            time_ago = f"{hours} hour{'s' if hours > 1 else ''} ago"
        elif time_diff.seconds > 60:
            minutes = time_diff.seconds // 60
            time_ago = f"{minutes} min ago"
        else:
            time_ago = "Just now"
        
        activities.append({
            "id": app.id,
            "icon": icon,
            "title": title,
            "meta": f"{app.program_type or 'Loan'} • ${app.loan_amount:,.0f}" if app.loan_amount else f"{app.program_type or 'Loan'}",
            "time": time_ago,
            "timestamp": app.created_at.isoformat()
        })
    
    return {"activities": activities}


@app.get("/applications")
async def list_applications(
    status: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """List applications for the organization"""
    
    query = db.query(LoanApplication).filter(
        LoanApplication.client_id == current_user.client_id
    )
    
    if status:
        query = query.filter(LoanApplication.status == status)
    
    total = query.count()
    apps = query.order_by(LoanApplication.created_at.desc()).offset(offset).limit(limit).all()
    
    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "applications": [{
            "id": a.id,
            "applicant_email": a.applicant_email,
            "property_address": a.property_address,
            "loan_amount": a.loan_amount,
            "decision": a.overall_decision,
            "status": a.status,
            "processing_time": a.processing_time_seconds,
            "created_at": a.created_at,
            "processed_by": a.processed_by.full_name if a.processed_by else None
        } for a in apps]
    }


# ==================== HELPER FUNCTIONS ====================

def _extract_with_regex(email_content: str) -> dict:
    """Simple regex extraction (fallback) - FIXED VERSION"""
    import re
    
    text = email_content
    extracted = {}
    
    # Extract units - must be followed by unit/units/doors (not just any number)
    units_match = re.search(r'(?:^|\n)\s*units?[:\s]+(\d+)', text, re.IGNORECASE)
    if not units_match:
        units_match = re.search(r'(\d+)\s*(?:unit|units|door|doors|plex)(?:\s|$|\n)', text, re.IGNORECASE)
    if units_match:
        extracted['units'] = int(units_match.group(1))
    
    # Extract address - look for "Address:" prefix or street patterns
    addr_match = re.search(r'address[:\s]+([^\n]+)', text, re.IGNORECASE)
    if addr_match:
        extracted['address'] = addr_match.group(1).strip()
    else:
        # Fallback to street pattern
        addr_match = re.search(r'(\d+\s+[^,\n]+?(?:Street|St|Avenue|Ave|Road|Rd|Blvd|Drive|Dr))', text, re.IGNORECASE)
        if addr_match:
            extracted['address'] = addr_match.group(1).strip()
    
    # Extract city - look for "City:" prefix
    city_match = re.search(r'city[:\s]+([^\n,]+)', text, re.IGNORECASE)
    if city_match:
        extracted['city'] = city_match.group(1).strip()
    else:
        # Fallback to pattern: City, ST
        city_match = re.search(r'([^,\n]+?),\s*[A-Z]{2}', text)
        if city_match:
            extracted['city'] = city_match.group(1).strip()
    
    # Extract state - look for "State:" prefix
    state_match = re.search(r'state[:\s]+([A-Z]{2})', text, re.IGNORECASE)
    if state_match:
        extracted['state'] = state_match.group(1).upper()
    else:
        # Fallback to pattern after city
        state_match = re.search(r',\s*([A-Z]{2})\s*\d{5}', text)
        if state_match:
            extracted['state'] = state_match.group(1).upper()
    
    # Extract zip - look for "Zip:" prefix first
    zip_match = re.search(r'zip(?:[_\s]?code)?:\s*\b(\d{5})\b', text, re.IGNORECASE)
    if zip_match:
        extracted['zip_code'] = zip_match.group(1)
    else:
        # Fallback to any 5-digit number that's not part of a larger number
        zip_match = re.search(r'\b(\d{5})\b(?!\d)', text)
        if zip_match:
            extracted['zip_code'] = zip_match.group(1)
    
    # Extract purchase price
    price_match = re.search(r'purchase[_\s]?price[:\s]+\$?([\d,]+)', text, re.IGNORECASE)
    if price_match:
        extracted['purchase_price'] = float(price_match.group(1).replace(',', ''))
    
    # Extract loan amount
    loan_match = re.search(r'loan[_\s]?amount[:\s]+\$?([\d,]+)', text, re.IGNORECASE)
    if loan_match:
        extracted['loan_amount'] = float(loan_match.group(1).replace(',', ''))
    
    # Extract estimated value (for refinances)
    value_match = re.search(r'estimated[_\s]?value[:\s]+\$?([\d,]+)', text, re.IGNORECASE)
    if value_match:
        extracted['estimated_value'] = float(value_match.group(1).replace(',', ''))
    
    # Extract credit score(s)
    # First try to find 3 scores
    three_scores = re.search(r'credit[_\s]?scores?[:\s]+(\d{3})[\s,]+(\d{3})[\s,]+(\d{3})', text, re.IGNORECASE)
    if three_scores:
        extracted['credit_score_1'] = int(three_scores.group(1))
        extracted['credit_score_2'] = int(three_scores.group(2))
        extracted['credit_score_3'] = int(three_scores.group(3))
    else:
        # Try to find single score
        single_score = re.search(r'credit[_\s]?scores?[:\s]+(\d{3})', text, re.IGNORECASE)
        if single_score:
            score = int(single_score.group(1))
            extracted['credit_score_1'] = score
            extracted['credit_score_2'] = score
            extracted['credit_score_3'] = score
    
    # Extract note type
    note_match = re.search(r'(30|15|20)\s*YR', text, re.IGNORECASE)
    if note_match:
        extracted['note_type'] = f"{note_match.group(1)} YR Fixed"
    else:
        extracted['note_type'] = "30 YR Fixed"  # Default
    
    # Extract points
    points_match = re.search(r'points?[:\s]+(\d+(?:\.\d+)?)', text, re.IGNORECASE)
    if points_match:
        extracted['points_to_lender'] = float(points_match.group(1))
    else:
        extracted['points_to_lender'] = 0.0
    
    # Set defaults
    extracted.setdefault('property_type', 'Multifamily')
    extracted.setdefault('asset_class', 'C')
    if 'estimated_value' not in extracted and 'purchase_price' in extracted:
        extracted['estimated_value'] = extracted['purchase_price']
    
    return extracted


# ==================== EXCEL EXPORT ====================

@app.get("/applications/{application_id}/export")
async def export_application_excel(
    application_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Export application as Excel file using the original sizer template
    """
    from fastapi.responses import FileResponse
    
    app = db.query(LoanApplication).filter(
        LoanApplication.id == application_id,
        LoanApplication.client_id == current_user.client_id
    ).first()
    
    if not app:
        raise HTTPException(404, "Application not found")
    
    # Check if output file exists
    if app.output_excel_path and Path(app.output_excel_path).exists():
        # Return existing processed file
        return FileResponse(
            app.output_excel_path,
            filename=f"loan_application_{application_id}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # Otherwise, generate from template
    template = db.query(ExcelTemplate).filter(
        ExcelTemplate.id == app.template_id
    ).first()
    
    if not template:
        raise HTTPException(400, "Template not found for this application")
    
    try:
        # Create loan application model from stored data
        loan_data = {
            'units': app.units or 1,
            'address': app.property_address or '',
            'city': app.property_city or '',
            'state': app.property_state or '',
            'zip_code': app.property_zip or '',
            'estimated_value': app.estimated_value or 0,
            'purchase_price': app.purchase_price or 0,
            'loan_amount': app.loan_amount or 0,
            'note_type': app.note_type or '30 Year Fixed',
            'credit_score_1': app.credit_score_1 or 700,
            'credit_score_2': app.credit_score_2 or 700,
            'credit_score_3': app.credit_score_3 or 700,
            'points_to_lender': app.points_to_lender or 0,
        }
        
        loan_app = LoanAppModel(**loan_data)
        
        # Process through sizer
        processor = SizerProcessor(template.file_path)
        result = processor.process_application(loan_app, app.interest_rate or 7.5)
        
        # Update record
        app.output_excel_path = result.output_file
        db.commit()
        
        return FileResponse(
            result.output_file,
            filename=f"loan_application_{application_id}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        raise HTTPException(500, f"Export failed: {str(e)}")


# ==================== SUPPORTED FORMATS ====================

@app.get("/supported-formats")
async def get_supported_formats():
    """Get list of supported file formats"""
    return {
        "formats": file_parser.SUPPORTED_FORMATS,
        "all_extensions": file_parser.get_supported_formats()
    }


# ==================== HEALTH CHECK ====================

@app.get("/health")
async def health_check():
    """Health check endpoint for Render"""
    return {
        "status": "healthy",
        "service": "loan-sizer-api",
        "version": "2.0.0",
        "timestamp": datetime.utcnow().isoformat()
    }

# ==================== INITIALIZATION ====================

@app.on_event("startup")
async def startup_event():
    """Initialize on startup"""
    init_database()
    print("🚀 Loan Sizer SaaS Platform v2.1.0 initialized")
    print(f"📁 Supported file formats: {', '.join(file_parser.get_supported_formats())}")
    print(f"✨ Features: Universal file parsing, Excel export, Real-time dashboard")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=5050)
